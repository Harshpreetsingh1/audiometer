#!/usr/bin/env python3
"""
PC Audiometer GUI - Simple and effective hearing test interface
Requires USB headphones connected to the PC
"""

import PySimpleGUI as sg
import threading
import sounddevice as sd
import os
import sys
from ascending_method import AscendingMethod

# Global test instance
current_test = None
test_thread = None
is_running = False
test_lock = threading.Lock()


def update_output(window, buffer, new_text):
    """Append new_text to buffer, update the Multiline and scroll to end."""
    if new_text is None:
        return buffer
    buffer += new_text
    try:
        window['-OUTPUT-'].update(buffer)
        widget = window['-OUTPUT-'].Widget
        widget.see('end')
    except Exception:
        try:
            window['-OUTPUT-'].update(buffer)
        except Exception:
            pass
    return buffer


def run_test_thread(device_id, window, subject_name=None):
    """
    Runs the hearing test in a background thread.
    This prevents the UI from freezing during the test.
    """
    global current_test, is_running
    
    is_running = True
    try:
        # Helper to forward printed output from the test thread to the GUI
        class WindowWriter:
            def __init__(self, win):
                self.win = win
            def write(self, s):
                if s and not s.isspace():
                    # Send small chunks to avoid flooding
                    try:
                        self.win.write_event_value('-TEST_LOG-', s)
                    except Exception:
                        pass
            def flush(self):
                pass

        # Create test instance with specific device
        test = AscendingMethod(device_id=device_id)
        # Provide UI window and subject name to controller for progress and final export
        try:
            test.ctrl.ui_window = window
            test.ctrl.subject_name = subject_name
        except Exception:
            pass
        with test_lock:
            current_test = test
        
        # Signal test started
        window.write_event_value('-TEST_STARTED-', '')
        
        # Redirect stdout/stderr from the test thread to the GUI
        old_stdout = sys.stdout
        old_stderr = sys.stderr
        writer = WindowWriter(window)
        sys.stdout = writer
        sys.stderr = writer
        try:
            # Run the actual test
            test.run()
        finally:
            # Restore std streams
            sys.stdout = old_stdout
            sys.stderr = old_stderr
        
        # Signal test completed
        window.write_event_value('-TEST_FINISHED-', '')

        # After completion, export CSV results to Excel with subject name as heading
        try:
            import openpyxl
            csv_path = os.path.join(test.ctrl.config.results_path, test.ctrl.config.filename)
            xlsx_path = os.path.splitext(csv_path)[0] + '.xlsx'
            wb = openpyxl.Workbook()
            # Use subject name as sheet title if possible
            sheet_title = subject_name if subject_name else 'Results'
            # Excel sheet titles have a max length of 31
            try:
                ws = wb.active
                ws.title = sheet_title[:31]
            except Exception:
                ws = wb.active
            # Write subject name as heading in first row
            if subject_name:
                ws.cell(row=1, column=1, value=subject_name)
                start_row = 2
            else:
                start_row = 1
            # Read CSV and append rows
            import csv as _csv
            with open(csv_path, 'r', newline='') as rf:
                reader = _csv.reader(rf)
                for r_idx, row in enumerate(reader, start=start_row):
                    for c_idx, val in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=val)
            wb.save(xlsx_path)
            # Inform UI
            try:
                window.write_event_value('-TEST_LOG-', f"Saved Excel: {xlsx_path}\n")
            except Exception:
                pass
        except Exception as e:
            try:
                window.write_event_value('-TEST_LOG-', f"Excel export failed: {e}\n")
            except Exception:
                pass
        
    except Exception as e:
        print(f"\nERROR in test: {e}")
        import traceback
        traceback.print_exc()
        window.write_event_value('-TEST_ERROR-', str(e))
    finally:
        is_running = False
        # Clean up resources if test object exists
        with test_lock:
            if current_test is not None:
                try:
                    if hasattr(current_test, 'ctrl') and current_test.ctrl is not None:
                        current_test.ctrl.__exit__()
                except Exception as cleanup_error:
                    print(f"Error during cleanup: {cleanup_error}")
            current_test = None


def main():
    sg.theme('DarkBlue3')
    sg.set_options(font=("Helvetica", 11))
    # Ensure that assignments inside this function refer to the module-level
    # variables rather than creating new local variables.
    global current_test, test_thread, is_running

    # 1. Detect audio devices
    try:
        devices = sd.query_devices()
    except Exception as e:
        sg.popup_error(f"Error querying audio devices: {e}")
        return

    device_list = []
    default_device = None
    
    for i, d in enumerate(devices):
        if d['max_output_channels'] > 0:
            device_str = f"{i}: {d['name']}"
            device_list.append(device_str)
            # Prefer USB devices
            if 'USB' in d['name'] and default_device is None:
                default_device = device_str

    if not device_list:
        sg.popup_error("No audio output devices found!")
        return

    # 2. Create UI layout
    layout = [
        [sg.Text("PC AUDIOMETER", font=("Helvetica", 24, "bold"), text_color='lightblue')],
        [sg.Text("Hearing Assessment System", font=("Helvetica", 12, "italic"), text_color='lightgreen')],
        
        [sg.HSeparator()],
        
        [sg.Text("Audio Device Selection:", font=("Helvetica", 11, "bold"))],
        [sg.Combo(device_list, default_value=default_device, key='-DEVICE-', 
                  size=(60, 1), readonly=True, background_color='lightgray', text_color='black')],
        
        [sg.HSeparator()],
        
        [sg.Text("Test Status:", font=("Helvetica", 11, "bold"))],
        [sg.Text("Ready to start.", key='-STATUS-', font=("Helvetica", 11), 
                 text_color='yellow', background_color='#1a1a1a', pad=(10, 5), size=(70, 1))],
        
        [sg.HSeparator()],
        
        [sg.Button("START TEST", key='-START-', size=(20, 2), 
                   button_color=('white', 'green'), font=("Helvetica", 12, "bold")),
         sg.Button("STOP TEST", key='-STOP-', size=(20, 2), 
                   button_color=('white', 'red'), font=("Helvetica", 12, "bold"), disabled=True)],
        
        [sg.Text("Patient Response Button:", font=("Helvetica", 11, "bold"))],
        [sg.Button("I HEAR THE TONE!", key='-RESPONSE-', size=(40, 3), 
                   button_color=('white', 'blue'), font=("Helvetica", 14, "bold"), 
                   disabled=True, pad=(10, 10))],
        
        [sg.HSeparator()],
        
        [sg.Text("Test Output:", font=("Helvetica", 11, "bold"))],
        [sg.Multiline(size=(80, 12), key='-OUTPUT-', disabled=True, 
                      background_color='black', text_color='#00FF00', 
                      font=("Courier", 10))],
        [sg.Text('Progress:'), sg.ProgressBar(max_value=100, orientation='h', size=(50, 20), key='-PROG-')],
        
        [sg.HSeparator()],
        
        [sg.Button("View Results", key='-VIEW-', size=(15, 1)),
         sg.Button("Exit", key='-EXIT-', size=(15, 1))],
    ]

    # 3. Create window
    window = sg.Window("PC Audiometer", layout, finalize=True, 
                      size=(900, 800), resizable=False)

    # 4. Bind mouse button press/release events
    response_btn = window['-RESPONSE-']
    response_btn.bind('<ButtonPress-1>', '_PRESSED')
    response_btn.bind('<ButtonRelease-1>', '_RELEASED')
    # Bind spacebar press/release to act like the response button
    try:
        window.bind('<KeyPress-space>', '-SPACE_PRESSED')
        window.bind('<KeyRelease-space>', '-SPACE_RELEASED')
    except Exception:
        # Some PySimpleGUI backends may not support window.bind; ignore if unavailable
        pass

    # 5. Event loop
    output_buffer = ""
    
    while True:
        try:
            event, values = window.read(timeout=100)
        except Exception as e:
            print(f"Error reading events: {e}")
            break

        # --- Window close ---
        if event == sg.WINDOW_CLOSED or event == '-EXIT-':
            if is_running:
                sg.popup_warning("Test in progress. Please stop the test first.")
                continue
            break

        # Handle None event (timeout)
        if event is None:
            continue

        # --- Start test ---
        if event == '-START-':
            device_str = values['-DEVICE-']
            if not device_str:
                sg.popup_error("Please select an audio device first!")
                continue

            device_id = int(device_str.split(':')[0])
            print(f"[UI] START TEST clicked with device_id={device_id}")

            # Update UI
            window['-START-'].update(disabled=True)
            window['-STOP-'].update(disabled=False)
            window['-RESPONSE-'].update(disabled=False)
            window['-DEVICE-'].update(disabled=True)
            window['-STATUS-'].update("Test Running... Listen for tones and press button when you hear them.", 
                                      text_color='lightgreen')
            output_buffer = "Test started. Listening for your responses...\n"
            output_buffer = update_output(window, output_buffer, "Test started. Listening for your responses...\n")

            # Ask for subject name
            subject_name = sg.popup_get_text('Enter subject name (for results heading):', 'Subject Name')
            if subject_name is None:
                # User cancelled - reset UI
                window['-START-'].update(disabled=False)
                window['-STOP-'].update(disabled=True)
                window['-RESPONSE-'].update(disabled=True)
                window['-DEVICE-'].update(disabled=False)
                window['-STATUS-'].update("Ready to start.", text_color='yellow')
                continue

            # Start test thread
            print("[UI] Creating test thread...")
            test_thread = threading.Thread(
                target=run_test_thread, 
                args=(device_id, window, subject_name), 
                daemon=True
            )
            test_thread.start()
            print("[UI] Test thread started")

        # --- Stop test ---
        elif event == '-STOP-':
            # Reset UI
            window['-START-'].update(disabled=False)
            window['-STOP-'].update(disabled=True)
            window['-RESPONSE-'].update(disabled=True)
            window['-DEVICE-'].update(disabled=False)
            window['-STATUS-'].update("Test stopped by user.", text_color='red')
            output_buffer = update_output(window, output_buffer, "\nTest stopped.\n")
            is_running = False

        # --- Patient response button press ---
        elif event == '-RESPONSE-_PRESSED':
            if is_running and current_test:
                with test_lock:
                    if current_test and hasattr(current_test.ctrl, '_rpd'):
                        current_test.ctrl._rpd.ui_button_pressed()

        # --- Spacebar pressed (acts like response button) ---
        elif event == '-SPACE_PRESSED':
            if is_running and current_test:
                with test_lock:
                    if current_test and hasattr(current_test.ctrl, '_rpd'):
                        current_test.ctrl._rpd.ui_button_pressed()

        # --- Patient response button release ---
        elif event == '-RESPONSE-_RELEASED':
            if is_running and current_test:
                with test_lock:
                    if current_test and hasattr(current_test.ctrl, '_rpd'):
                        current_test.ctrl._rpd.ui_button_released()

        # --- Spacebar released (acts like response button) ---
        elif event == '-SPACE_RELEASED':
            if is_running and current_test:
                with test_lock:
                    if current_test and hasattr(current_test.ctrl, '_rpd'):
                        current_test.ctrl._rpd.ui_button_released()

        # --- Test events ---
        elif event == '-TEST_STARTED-':
            output_buffer = update_output(window, output_buffer, "Test initialization complete.\n")

        # --- Test log lines forwarded from test thread ---
        elif event == '-TEST_LOG-':
            # values[event] contains the printed text chunk
            log_text = values.get(event, '')
            # Append to buffer and update UI
            output_buffer = update_output(window, output_buffer, log_text)
            # If the test is asking the user to click, make the prompt explicit
            lt = log_text.lower()
            if 'click once' in lt or 'press the button' in lt or 'click the button' in lt or 'press button' in lt:
                try:
                    window['-STATUS-'].update("Please click the large blue 'I HEAR THE TONE!' button now", text_color='lightgreen')
                    window['-RESPONSE-'].update(disabled=False)
                    # Try to focus the response button (may work on Tk)
                    try:
                        window['-RESPONSE-'].Widget.focus_set()
                    except Exception:
                        pass
                except Exception:
                    pass

        # --- Progress update ---
        elif event == '-PROGRESS-':
            try:
                percent = int(values.get(event) or 0)
            except Exception:
                percent = 0
            try:
                window['-PROG-'].update(percent)
                window['-STATUS-'].update(f"Test Running... {percent}%", text_color='lightgreen')
            except Exception:
                pass

        elif event == '-TEST_FINISHED-':
            window['-STATUS-'].update("Test Completed!", text_color='lightgreen')
            output_buffer = update_output(window, output_buffer, "\n✓ Test finished! Results have been saved.\n")
            window['-START-'].update(disabled=False)
            window['-STOP-'].update(disabled=True)
            window['-RESPONSE-'].update(disabled=True)
            window['-DEVICE-'].update(disabled=False)
            sg.popup_ok("Hearing test completed!\n\nResults saved to audiometer/results/", 
                       title="Test Complete")

        elif event == '-TEST_ERROR-':
            error = values.get(event, "Unknown error")
            window['-STATUS-'].update(f"Error: {error}", text_color='red')
            output_buffer = update_output(window, output_buffer, f"\n✗ ERROR: {error}\n")
            window['-START-'].update(disabled=False)
            window['-STOP-'].update(disabled=True)
            window['-RESPONSE-'].update(disabled=True)
            window['-DEVICE-'].update(disabled=False)
            sg.popup_error(f"Test Error: {error}")

        # --- View results ---
        elif event == '-VIEW-':
            results_dir = os.path.join(os.getcwd(), 'audiometer', 'results')
            if os.path.exists(results_dir) and os.listdir(results_dir):
                try:
                    if sys.platform == 'win32':
                        os.startfile(results_dir)
                    elif sys.platform == 'darwin':
                        os.system(f'open "{results_dir}"')
                    else:
                        os.system(f'xdg-open "{results_dir}"')
                except Exception as e:
                    sg.popup_error(f"Could not open results folder: {e}")
            else:
                sg.popup_warning("No results found. Run a test first.")

    window.close()


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"Fatal error in main: {e}")
        import traceback
        traceback.print_exc()
        sg.popup_error(f"Fatal application error:\n{e}\n\nCheck console for details.", title="Error")
